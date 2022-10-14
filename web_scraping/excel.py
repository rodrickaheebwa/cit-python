import openpyxl

excel = openpyxl.Workbook()
# print(excel.sheetnames)

# work with a sheet
sheet = excel.active
sheet.title = 'New Sheet Title'
# print(excel.sheetnames)

# add a row of data
sheet.append('col_1_value', 'col_2_value', 'col_3_value')

# save excel to create it
# excel.save('excel_workbook.xlsx')